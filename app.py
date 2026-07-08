from flask import Flask, request
from flask_cors import CORS
from flask_restx import Api, Resource, fields
import requests
import os
from dotenv import load_dotenv
import sqlite3
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import io


# Load environment variables from .env file
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'default_secret_key')

# Enable CORS for React Native access
CORS(app)

# Initialize Flask-RESTx API with Swagger documentation
api = Api(
    app,
    version='3.0',
    title='Multi-Product Stock Management API',
    description='''## API Documentation for Multi-Product Stock Management System

**Version:** 3.0 - Hybrid JSON Architecture

### Features:
- **Dynamic Product Management** - Support multiple product types with flexible fields
- **Product Type Configuration** - Define and manage product types dynamically
- **Hybrid JSON Storage** - Native columns + flexible JSON fields
- **Advanced Filtering** - Search in both native columns and JSON fields
- **Backward Compatible** - Existing phone API still works

### Architecture:
- **Native Fields**: `product_type`, `phone_number`, `operator`, `location`, `status`
- **Flexible Fields**: Any custom fields stored as JSON in `flex_fields`
- **Auto Detection**: System automatically separates native vs flex fields

### Quick Start:
1. Register product type via `/product-types`
2. Add products via `/products` with any custom fields
3. Search/filter using query params (including JSON fields)

**Base URL:** `/api`
**Swagger UI:** `/api/docs`''',
    doc='/',
    prefix='/api'
)

# Define Namespaces
product_types_ns = api.namespace('product-types', description='Product Type Configuration Operations')
products_ns = api.namespace('products', description='Dynamic Product Operations (Hybrid JSON)')
phone_ns = api.namespace('phone', description='Phone Number Operations (Legacy)')

# Configuration
API_KEY_PROVIDER = os.getenv('API_KEY_PROVIDER')
FLASK_ENV = os.getenv('FLASK_ENV', 'development')
DATABASE = 'stok_kartu.db'

# Native fields that exist as columns in database
NATIVE_FIELDS = {'phone_number', 'operator', 'location', 'status', 'product_type'}


def get_db_connection():
    """Create database connection"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    # Enable foreign keys
    conn.execute('PRAGMA foreign_keys = ON')
    return conn


def init_db():
    """Initialize database with required tables and migrations"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. Create product_type_config table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS product_type_config (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type_code TEXT UNIQUE NOT NULL,
            type_name TEXT NOT NULL,
            required_fields TEXT NOT NULL,
            optional_fields TEXT,
            description TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 2. Add product_type column to products table (if not exists)
    cursor.execute('''
        SELECT count(*) FROM pragma_table_info('products') 
        WHERE name='product_type'
    ''')
    if cursor.fetchone()[0] == 0:
        cursor.execute('ALTER TABLE products ADD COLUMN product_type TEXT DEFAULT "KARTU"')
    
    # 3. Add flex_fields column to products table (if not exists)
    cursor.execute('''
        SELECT count(*) FROM pragma_table_info('products') 
        WHERE name='flex_fields'
    ''')
    if cursor.fetchone()[0] == 0:
        cursor.execute('ALTER TABLE products ADD COLUMN flex_fields TEXT DEFAULT NULL')
    
    # 4. Add notes column to products table (if not exists)
    cursor.execute('''
        SELECT count(*) FROM pragma_table_info('products') 
        WHERE name='notes'
    ''')
    if cursor.fetchone()[0] == 0:
        cursor.execute('ALTER TABLE products ADD COLUMN notes TEXT DEFAULT NULL')
    
    # 5. Create index for product_type
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_product_type 
        ON products(product_type)
    ''')
    
    # 6. Insert default product type config for Kartu
    cursor.execute('''
        INSERT OR IGNORE INTO product_type_config 
        (type_code, type_name, required_fields, optional_fields, description)
        VALUES (?, ?, ?, ?, ?)
    ''', (
        'KARTU',
        'Kartu Perdana',
        json.dumps(['phone_number', 'operator']),
        json.dumps(['harga_modal', 'masa_aktif', 'bonus', 'catatan']),
        'Produk kartu perdana fisik dengan berbagai operator'
    ))
    
    conn.commit()
    conn.close()


# Initialize database on startup
init_db()


# ==========================================
# API Models for Swagger Documentation
# ==========================================

# Product Type Config Models
product_type_config_model = api.model('ProductTypeConfig', {
    'type_code': fields.String(required=True, description='Product type code (e.g., KARTU, TOKEN, WIFI)'),
    'type_name': fields.String(required=True, description='Product type name'),
    'required_fields': fields.List(fields.String, required=True, description='List of required field names'),
    'optional_fields': fields.List(fields.String, description='List of optional field names'),
    'description': fields.String(description='Product type description'),
    'is_active': fields.Boolean(description='Is this product type active?')
})

product_type_update_model = api.model('ProductTypeUpdate', {
    'type_name': fields.String(description='Product type name'),
    'required_fields': fields.List(fields.String, description='List of required field names'),
    'optional_fields': fields.List(fields.String, description='List of optional field names'),
    'description': fields.String(description='Product type description'),
    'is_active': fields.Boolean(description='Is this product type active?')
})

# Dynamic Product Models
dynamic_product_model = api.model('DynamicProduct', {
    'product_type': fields.String(description='Product type code (auto-defaults to KARTU)'),
    'phone_number': fields.String(description='Phone number (native field)'),
    'operator': fields.String(description='Operator name (native field)'),
    'location': fields.String(description='Location (native field)'),
    'status': fields.String(description='Status: active/inactive (native field)'),
    'notes': fields.String(description='General notes'),
    'any_custom_field': fields.String(description='Any additional custom fields will be stored in flex_fields JSON')
})

product_update_model = api.model('ProductUpdate', {
    'phone_number': fields.String(description='Phone number'),
    'operator': fields.String(description='Operator name'),
    'location': fields.String(description='Location'),
    'status': fields.String(description='Status: active/inactive'),
    'notes': fields.String(description='General notes'),
    'any_custom_field': fields.String(description='Any other field will be stored in flex_fields JSON')
})

# Legacy Phone Models (keep existing)
phone_add_model = api.model('PhoneAdd', {
    'phone_number': fields.String(required=True, description='Phone Number'),
    'operator': fields.String(required=True, description='Operator'),
    'location': fields.String(required=True, description='Location')
})

phone_bulk_add_model = api.model('PhoneBulkAdd', {
    'phone_numbers': fields.List(fields.String, required=True, description='List of Phone Numbers'),
    'operator': fields.String(required=True, description='Operator'),
    'location': fields.String(required=True, description='Location')
})

product_check_model = api.model('ProductCheck', {
    'phone_number': fields.String(required=True, description='Phone Number to Check')
})


# ==========================================
# PRODUCT TYPE CONFIG API
# ==========================================

@product_types_ns.route('')
class ProductTypeList(Resource):
    """Manage product type configurations"""
    
    @product_types_ns.marshal_list_with(product_type_config_model)
    @product_types_ns.response(200, 'Product types retrieved successfully')
    def get(self):
        """Get all product type configurations"""
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            show_inactive = request.args.get('show_inactive', 'false').lower() == 'true'
            
            query = 'SELECT * FROM product_type_config WHERE 1=1'
            params = []
            
            if not show_inactive:
                query += ' AND is_active = 1'
            
            query += ' ORDER BY type_name ASC'
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            conn.close()
            
            result = []
            for row in rows:
                result.append({
                    'id': row['id'],
                    'type_code': row['type_code'],
                    'type_name': row['type_name'],
                    'required_fields': json.loads(row['required_fields']) if row['required_fields'] else [],
                    'optional_fields': json.loads(row['optional_fields']) if row['optional_fields'] else [],
                    'description': row['description'],
                    'is_active': bool(row['is_active']),
                    'created_at': row['created_at']
                })
            
            return result, 200
            
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500
    
    @product_types_ns.expect(product_type_config_model)
    @product_types_ns.response(201, 'Product type created successfully')
    @product_types_ns.response(400, 'Invalid input')
    @product_types_ns.response(409, 'Product type already exists')
    def post(self):
        """Register a new product type configuration"""
        try:
            data = request.get_json()
            
            type_code = data.get('type_code', '').strip().upper()
            type_name = data.get('type_name', '').strip()
            required_fields = data.get('required_fields', [])
            optional_fields = data.get('optional_fields', [])
            description = data.get('description', '').strip()
            is_active = data.get('is_active', 1)
            
            if not type_code or not type_name:
                return {'status': 'error', 'message': 'type_code and type_name are required'}, 400
            
            if not isinstance(required_fields, list):
                return {'status': 'error', 'message': 'required_fields must be an array'}, 400
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            query = '''
                INSERT INTO product_type_config 
                (type_code, type_name, required_fields, optional_fields, description, is_active)
                VALUES (?, ?, ?, ?, ?, ?)
            '''
            cursor.execute(query, (
                type_code,
                type_name,
                json.dumps(required_fields),
                json.dumps(optional_fields) if optional_fields else None,
                description,
                1 if is_active else 0
            ))
            
            conn.commit()
            config_id = cursor.lastrowid
            conn.close()
            
            return {
                'status': 'success',
                'data': {
                    'id': config_id,
                    'type_code': type_code,
                    'type_name': type_name,
                    'required_fields': required_fields,
                    'optional_fields': optional_fields,
                    'description': description
                },
                'message': f'Product type {type_name} registered successfully'
            }, 201
            
        except sqlite3.IntegrityError:
            return {'status': 'error', 'message': 'Product type already exists'}, 409
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500


@product_types_ns.route('/<string:type_code>')
class ProductTypeDetail(Resource):
    """Get, update or delete a specific product type"""
    
    @product_types_ns.marshal_with(product_type_config_model)
    @product_types_ns.response(200, 'Product type retrieved')
    @product_types_ns.response(404, 'Product type not found')
    def get(self, type_code):
        """Get a specific product type configuration"""
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM product_type_config WHERE type_code = ?', (type_code.upper(),))
            row = cursor.fetchone()
            conn.close()
            
            if not row:
                return {'status': 'error', 'message': 'Product type not found'}, 404
            
            result = {
                'id': row['id'],
                'type_code': row['type_code'],
                'type_name': row['type_name'],
                'required_fields': json.loads(row['required_fields']) if row['required_fields'] else [],
                'optional_fields': json.loads(row['optional_fields']) if row['optional_fields'] else [],
                'description': row['description'],
                'is_active': bool(row['is_active']),
                'created_at': row['created_at']
            }
            
            return result, 200
            
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500
    
    @product_types_ns.expect(product_type_update_model)
    @product_types_ns.response(200, 'Product type updated successfully')
    @product_types_ns.response(404, 'Product type not found')
    def put(self, type_code):
        """Update product type configuration"""
        try:
            data = request.get_json()
            
            if not data:
                return {'status': 'error', 'message': 'No data provided'}, 400
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT id FROM product_type_config WHERE type_code = ?', (type_code.upper(),))
            if not cursor.fetchone():
                conn.close()
                return {'status': 'error', 'message': 'Product type not found'}, 404
            
            allowed_fields = ['type_name', 'required_fields', 'optional_fields', 'description', 'is_active']
            update_fields = []
            params = []
            
            for field in allowed_fields:
                if field in data:
                    update_fields.append(f'{field} = ?')
                    value = data[field]
                    
                    if field in ['required_fields', 'optional_fields'] and isinstance(value, list):
                        value = json.dumps(value)
                    elif field == 'is_active':
                        value = 1 if value else 0
                    
                    params.append(value)
            
            if not update_fields:
                conn.close()
                return {'status': 'error', 'message': 'No valid fields to update'}, 400
            
            params.append(type_code.upper())
            query = f'UPDATE product_type_config SET {", ".join(update_fields)} WHERE type_code = ?'
            cursor.execute(query, params)
            
            conn.commit()
            conn.close()
            
            return {'status': 'success', 'message': 'Product type updated successfully'}, 200
            
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500
    
    @product_types_ns.response(200, 'Product type deleted successfully')
    @product_types_ns.response(404, 'Product type not found')
    def delete(self, type_code):
        """Delete a product type configuration"""
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT id FROM product_type_config WHERE type_code = ?', (type_code.upper(),))
            if not cursor.fetchone():
                conn.close()
                return {'status': 'error', 'message': 'Product type not found'}, 404
            
            cursor.execute('DELETE FROM product_type_config WHERE type_code = ?', (type_code.upper(),))
            conn.commit()
            conn.close()
            
            return {'status': 'success', 'message': 'Product type deleted successfully'}, 200
            
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500


# Health check endpoint (keep as simple Flask route)
@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    from flask import jsonify
    return jsonify({
        'status': 'success',
        'message': 'API is running',
        'environment': FLASK_ENV,
        'docs': '/api/docs'
    }), 200


# Product Check Endpoint (Legacy - for backward compatibility)
@phone_ns.route('/add')
class PhoneAdd(Resource):
    """Add single phone number to database (Legacy)"""
    
    @phone_ns.expect(phone_add_model)
    @phone_ns.response(201, 'Phone number added successfully')
    @phone_ns.response(400, 'Invalid input')
    @phone_ns.response(409, 'Phone number already exists')
    def post(self):
        """Add a new phone number manually to database"""
        data = request.get_json()
        
        if not data or 'phone_number' not in data:
            return {'status': 'error', 'message': 'phone_number is required'}, 400
        
        phone_number = data['phone_number'].strip()
        operator = data.get('operator', '').strip().lower()
        location = data.get('location', '').strip()
        status = data.get('status', 'active')
        
        if not phone_number:
            return {'status': 'error', 'message': 'phone_number cannot be empty'}, 400
        
        if not operator:
            return {'status': 'error', 'message': 'operator is required'}, 400
        
        if not location:
            return {'status': 'error', 'message': 'location is required'}, 400
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT id FROM products WHERE phone_number = ?', (phone_number,))
            existing = cursor.fetchone()
            
            if existing:
                conn.close()
                return {
                    'status': 'error',
                    'message': 'Nomor sudah ada dalam database',
                    'data': {'phone_number': phone_number, 'exists': True}
                }, 409
            
            cursor.execute('''
                INSERT INTO products (phone_number, operator, location, status, product_type)
                VALUES (?, ?, ?, ?, 'KARTU')
            ''', (phone_number, operator, location, status))
            
            conn.commit()
            product_id = cursor.lastrowid
            conn.close()
            
            return {
                'status': 'success',
                'data': {
                    'id': product_id,
                    'phone_number': phone_number,
                    'operator': operator,
                    'location': location,
                    'status': status,
                    'product_type': 'KARTU',
                    'exists': True
                },
                'message': 'Nomor berhasil ditambahkan ke database'
            }, 201
            
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500


@phone_ns.route('/add-bulk')
class PhoneBulkAdd(Resource):
    """Bulk add phone numbers to database (Legacy)"""
    
    @phone_ns.expect(phone_bulk_add_model)
    @phone_ns.response(201, 'Phone numbers added successfully')
    @phone_ns.response(400, 'Invalid input')
    def post(self):
        """Add multiple phone numbers at once"""
        data = request.get_json()
        
        if not data or 'phone_numbers' not in data:
            return {'status': 'error', 'message': 'phone_numbers (array) is required'}, 400
        
        phone_numbers = data['phone_numbers']
        
        if not isinstance(phone_numbers, list) or len(phone_numbers) == 0:
            return {'status': 'error', 'message': 'phone_numbers must be a non-empty array'}, 400
        
        operator = data.get('operator', '').strip().lower()
        location = data.get('location', '').strip()
        
        if not operator:
            return {'status': 'error', 'message': 'operator is required'}, 400
        
        if not location:
            return {'status': 'error', 'message': 'location is required'}, 400
        
        added = []
        duplicates = []
        errors = []
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            for phone_number in phone_numbers:
                phone_number = phone_number.strip()
                
                if not phone_number:
                    continue
                
                try:
                    cursor.execute('SELECT id FROM products WHERE phone_number = ?', (phone_number,))
                    existing = cursor.fetchone()
                    
                    if existing:
                        duplicates.append(phone_number)
                        continue
                    
                    cursor.execute('''
                        INSERT INTO products (phone_number, operator, location, status, product_type)
                        VALUES (?, ?, ?, 'active', 'KARTU')
                    ''', (phone_number, operator, location))
                    
                    added.append(phone_number)
                    
                except Exception as e:
                    errors.append({'phone_number': phone_number, 'error': str(e)})
            
            conn.commit()
            conn.close()
            
            return {
                'status': 'success',
                'data': {
                    'added': added,
                    'duplicates': duplicates,
                    'errors': errors,
                    'total_added': len(added),
                    'total_duplicates': len(duplicates),
                    'total_errors': len(errors)
                },
                'message': f'Berhasil menambahkan {len(added)} nomor'
            }, 201
            
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500


@phone_ns.route('/check-bulk')
class PhoneBulkCheck(Resource):
    """Bulk check phone numbers from Excel file"""
    
    @phone_ns.response(200, 'Check completed')
    @phone_ns.response(400, 'Invalid input')
    def post(self):
        """Check multiple phone numbers from Excel file"""
        try:
            if 'file' not in request.files:
                return {'status': 'error', 'message': 'No file provided'}, 400
            
            file = request.files['file']
            
            if file.filename == '':
                return {'status': 'error', 'message': 'No file selected'}, 400
            
            if not file.filename.endswith(('.xlsx', '.xls')):
                return {'status': 'error', 'message': 'File must be Excel format (.xlsx or .xls)'}, 400
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            phone_numbers = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    phone_numbers.append(str(row[0]).strip())
            
            results = []
            found_count = 0
            not_found_count = 0
            
            for phone_number in phone_numbers:
                if not phone_number:
                    continue
                
                cursor.execute('''
                    SELECT id, phone_number, operator, location, status, product_type, created_at
                    FROM products
                    WHERE phone_number = ?
                ''', (phone_number,))
                
                product = cursor.fetchone()
                
                if product:
                    found_count += 1
                    results.append({
                        'phone_number': phone_number,
                        'exists': True,
                        'product': {
                            'id': product['id'],
                            'phone_number': product['phone_number'],
                            'operator': product['operator'],
                            'location': product['location'],
                            'status': product['status'],
                            'product_type': product['product_type'],
                            'created_at': product['created_at']
                        }
                    })
                else:
                    not_found_count += 1
                    results.append({
                        'phone_number': phone_number,
                        'exists': False,
                        'product': None
                    })
            
            conn.close()
            
            return {
                'status': 'success',
                'data': {
                    'results': results,
                    'total_checked': len(results),
                    'found': found_count,
                    'not_found': not_found_count
                },
                'message': f'Checked {len(results)} numbers: {found_count} found, {not_found_count} not found'
            }, 200
            
        except Exception as e:
            return {'status': 'error', 'message': f'Bulk check failed: {str(e)}'}, 500


@products_ns.route('/check')
class ProductCheck(Resource):
    """Check if phone number exists in database (Legacy)"""
    
    @products_ns.expect(product_check_model)
    @products_ns.response(200, 'Check completed')
    @products_ns.response(400, 'Invalid input')
    def post(self):
        """Check if phone number exists and get product info"""
        data = request.get_json()
        
        if not data or 'phone_number' not in data:
            return {'status': 'error', 'message': 'phone_number is required'}, 400
        
        phone_number = data['phone_number'].strip()
        
        if not phone_number:
            return {'status': 'error', 'message': 'phone_number cannot be empty'}, 400
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT id, phone_number, operator, location, status, product_type, created_at
                FROM products
                WHERE phone_number = ?
            ''', (phone_number,))
            
            product = cursor.fetchone()
            conn.close()
            
            if product:
                return {
                    'status': 'success',
                    'data': {
                        'exists': True,
                        'product': {
                            'id': product['id'],
                            'phone_number': product['phone_number'],
                            'operator': product['operator'],
                            'location': product['location'],
                            'status': product['status'],
                            'product_type': product['product_type'],
                            'created_at': product['created_at']
                        }
                    },
                    'message': 'Nomor ditemukan dalam database'
                }, 200
            else:
                return {
                    'status': 'success',
                    'data': {
                        'exists': False,
                        'product': None
                    },
                    'message': 'Nomor tidak ditemukan dalam database'
                }, 200
                
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500




# ==========================================
# DYNAMIC PRODUCTS API (HYBRID JSON)
# ==========================================

@products_ns.route('')
class DynamicProductList(Resource):
    """Get all products or add new product with dynamic fields"""
    
    @products_ns.response(200, 'Products retrieved successfully')
    def get(self):
        """Get all products with optional filtering (including JSON fields)"""
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            base_query = 'SELECT * FROM products WHERE 1=1'
            query_params = []
            
            # Process all query parameters for filtering
            for key, value in request.args.items():
                if key in ['search', 'page', 'limit', 'sort_by', 'sort_order']:
                    continue
                
                if key in NATIVE_FIELDS:
                    base_query += f' AND {key} = ?'
                    query_params.append(value)
                else:
                    # Search inside JSON flex_fields
                    base_query += " AND json_extract(flex_fields, '$.\"" + key + "\"') = ?"
                    query_params.append(value)
            
            # Global search across all fields
            search = request.args.get('search')
            if search:
                search_pattern = f'%{search}%'
                base_query += '''
                    AND (
                        phone_number LIKE ? 
                        OR operator LIKE ? 
                        OR location LIKE ?
                        OR notes LIKE ?
                        OR flex_fields LIKE ?
                    )
                '''
                query_params.extend([search_pattern] * 5)
            
            # Sorting
            sort_by = request.args.get('sort_by', 'created_at')
            sort_order = request.args.get('sort_order', 'DESC')
            
            if sort_by in ['created_at', 'updated_at', 'phone_number', 'operator', 'location', 'status']:
                base_query += f' ORDER BY {sort_by} {sort_order}'
            else:
                base_query += ' ORDER BY created_at DESC'
            
            cursor.execute(base_query, query_params)
            rows = cursor.fetchall()
            conn.close()
            
            # Flatten JSON fields for response
            cleaned_results = []
            for row in rows:
                flat_item = {}
                
                # Add native fields
                for key in row.keys():
                    if key != 'flex_fields' and row[key] is not None:
                        flat_item[key] = row[key]
                
                # Merge flex_fields into flat structure
                if row['flex_fields']:
                    try:
                        flex_dict = json.loads(row['flex_fields'])
                        flat_item.update(flex_dict)
                    except:
                        pass
                
                cleaned_results.append(flat_item)
            
            return cleaned_results, 200
            
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500
    
    @products_ns.expect(dynamic_product_model)
    @products_ns.response(201, 'Product created successfully')
    @products_ns.response(400, 'Invalid input')
    def post(self):
        """Add new product with dynamic/flexible fields"""
        try:
            data = request.get_json()
            
            if not data:
                return {'status': 'error', 'message': 'Request body is required'}, 400
            
            # Get product type (default to KARTU)
            p_type = data.get('product_type', 'KARTU').strip().upper()
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Validate against product_type_config if exists
            cursor.execute('SELECT required_fields FROM product_type_config WHERE type_code = ?', (p_type,))
            config = cursor.fetchone()
            
            if config and config['required_fields']:
                req_fields = json.loads(config['required_fields'])
                for field in req_fields:
                    if field not in data or not data[field]:
                        conn.close()
                        return {
                            'status': 'error',
                            'message': f"Field '{field}' is required for product type {p_type}"
                        }, 400
            
            # Separate native fields vs flex fields
            native_data = {}
            flex_data = {}
            
            for key, value in data.items():
                if key in NATIVE_FIELDS:
                    native_data[key] = value
                else:
                    flex_data[key] = value
            
            # Set default product_type
            if 'product_type' not in native_data:
                native_data['product_type'] = p_type
            
            # Set default status
            if 'status' not in native_data or not native_data['status']:
                native_data['status'] = 'active'
            
            # Convert flex_data to JSON string
            flex_fields_json = json.dumps(flex_data) if flex_data else None
            
            # Insert into database
            query = '''
                INSERT INTO products 
                (product_type, phone_number, operator, location, status, flex_fields)
                VALUES (?, ?, ?, ?, ?, ?)
            '''
            values = (
                native_data.get('product_type'),
                native_data.get('phone_number'),
                native_data.get('operator'),
                native_data.get('location'),
                native_data.get('status'),
                flex_fields_json
            )
            
            cursor.execute(query, values)
            conn.commit()
            new_id = cursor.lastrowid
            conn.close()
            
            return {
                'status': 'success',
                'data': {
                    'id': new_id,
                    'product_type': native_data.get('product_type'),
                    **native_data,
                    **flex_data
                },
                'message': 'Product created successfully'
            }, 201
            
        except sqlite3.IntegrityError as e:
            return {'status': 'error', 'message': f'Duplicate entry: {str(e)}'}, 409
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500


@products_ns.route('/<int:product_id>')
class DynamicProductDetail(Resource):
    """Get, update or delete a specific product"""
    
    @products_ns.response(200, 'Product retrieved')
    @products_ns.response(404, 'Product not found')
    def get(self, product_id):
        """Get a specific product by ID"""
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM products WHERE id = ?', (product_id,))
            row = cursor.fetchone()
            conn.close()
            
            if not row:
                return {'status': 'error', 'message': 'Product not found'}, 404
            
            # Flatten the response
            flat_item = {}
            for key in row.keys():
                if key != 'flex_fields' and row[key] is not None:
                    flat_item[key] = row[key]
            
            if row['flex_fields']:
                try:
                    flex_dict = json.loads(row['flex_fields'])
                    flat_item.update(flex_dict)
                except:
                    pass
            
            return flat_item, 200
            
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500
    
    @products_ns.expect(product_update_model)
    @products_ns.response(200, 'Product updated successfully')
    @products_ns.response(404, 'Product not found')
    def put(self, product_id):
        """Update product information (with flex_fields merge)"""
        try:
            data = request.get_json()
            
            if not data:
                return {'status': 'error', 'message': 'No data provided'}, 400
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Check if product exists
            cursor.execute('SELECT * FROM products WHERE id = ?', (product_id,))
            existing = cursor.fetchone()
            
            if not existing:
                conn.close()
                return {'status': 'error', 'message': 'Product not found'}, 404
            
            # Separate native fields vs flex fields
            native_updates = {}
            flex_updates = {}
            
            for key, value in data.items():
                if key in NATIVE_FIELDS:
                    native_updates[key] = value
                else:
                    flex_updates[key] = value
            
            # Update native fields
            if native_updates:
                update_fields = []
                params = []
                
                for field, value in native_updates.items():
                    update_fields.append(f'{field} = ?')
                    params.append(value)
                
                update_fields.append('updated_at = CURRENT_TIMESTAMP')
                params.append(product_id)
                
                query = f'UPDATE products SET {", ".join(update_fields)} WHERE id = ?'
                cursor.execute(query, params)
            
            # Update flex_fields (MERGE strategy)
            if flex_updates:
                existing_flex = {}
                if existing['flex_fields']:
                    try:
                        existing_flex = json.loads(existing['flex_fields'])
                    except:
                        pass
                
                # Merge old and new flex data
                existing_flex.update(flex_updates)
                new_flex_json = json.dumps(existing_flex)
                
                cursor.execute(
                    'UPDATE products SET flex_fields = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                    (new_flex_json, product_id)
                )
            
            conn.commit()
            conn.close()
            
            return {'status': 'success', 'message': 'Product updated successfully'}, 200
            
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500
    
    @products_ns.response(200, 'Product deleted successfully')
    @products_ns.response(404, 'Product not found')
    def delete(self, product_id):
        """Delete a product from database"""
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT id FROM products WHERE id = ?', (product_id,))
            if not cursor.fetchone():
                conn.close()
                return {'status': 'error', 'message': 'Product not found'}, 404
            
            cursor.execute('DELETE FROM products WHERE id = ?', (product_id,))
            conn.commit()
            conn.close()
            
            return {'status': 'success', 'message': 'Product deleted successfully'}, 200
            
        except Exception as e:
            return {'status': 'error', 'message': f'Database error: {str(e)}'}, 500


if __name__ == '__main__':
    app.run(
        host='0.0.0.0',
        port=5008,
        debug=(FLASK_ENV == 'development')
    )